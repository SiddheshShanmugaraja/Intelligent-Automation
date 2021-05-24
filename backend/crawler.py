import json, os, re
import pandas as pd
import openpyxl as op
from bs4 import BeautifulSoup
from selenium import webdriver
from urllib.parse import urlparse
from urllib.request import urlopen, Request
from selenium.webdriver.chrome.options import Options

with open("backend/config.json", "r") as f:
    config = json.load(f)

EXCEL_DIR = config.get("EXCEL_DIR") 
CHROME_DRIVER = config.get("CHROME_DRIVER")
EXCEL_EXTENSION = ".xlsx"
CSV_EXTENSION = ".csv"

get_filename = lambda x: re.search(r'https://(.*?).com', x).group(1) if re.search(r'https://(.*?).com', x) != None else re.search(r'http://(.*?):', x).group(1)

def crawl_urls(url_list, crawled_urls, driver, url):
    """Get a set of urls and crawl each url recursively

    Args:
        url_list ([type]): [description]
        crawled_urls ([type]): [description]
        driver ([type]): [description]
        url ([type]): [description]

    Returns:
        [type]: [description]
    """
    # Once the url is parsed, add it to crawled url list
    crawled_urls.append(url)
    print("url",url)
    driver.get(url)
    
    html = driver.page_source.encode("utf-8")

    soup = BeautifulSoup(html, features="html.parser")

    urls = soup.findAll("a")

    # Even if the url is not part of the same domain, it is still collected
    # But those urls not in the same domain are not parsed
    for a in urls:
        if (a.get("href")) and (a.get("href") not in url_list):
            page = a.get("href")
            page = url + page if page.count('/') == 1 else page
            if (domain in page) and (page not in crawled_urls):
                url_list.append(page)
    # Once all urls are crawled return the list to calling function
    return crawled_urls, url_list

def get_page_title(url):
    try:
        url = Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        soup = BeautifulSoup(urlopen(url), features="html.parser")
        return soup.title.get_text()
    except Exception as e:
        print(f"Error: {e}")
        return 'Unnamed'

def load_to_excel(lst):
    """Load the list into excel file using pandas

    Args:
        lst ([type]): [description]
    """
    # Load list to dataframe
    df = pd.DataFrame(lst)
    df.index += 1  # So that the excel column starts from 1

    # Write dataframe to excel
    print("\n\nXL Name:", xl_name, "\n\n")
    xlw = pd.ExcelWriter(xl_name)
    df.to_excel(xlw, sheet_name=sheet_name, index_label="#", header=["URL"])
    xlw.save()


def format_excel(xl, sheet="Sheet1"):
    """Get the excel file path and format the file
       If no sheet name is passed, by default take Sheet1

    Args:
        xl ([type]): [description]
        sheet (str, optional): [description]. Defaults to "Sheet1".
    """
    # Open the excel file
    wb = op.load_workbook(xl)
    ws = wb[sheet]

    # Freeze panes
    ws.freeze_panes = "B2"

    # Adjust column width
    cols = ("A", "B")
    widths = (5, 80)

    for combo in zip(cols, widths):
        ws.column_dimensions[combo[0]].width = combo[1]

    # define color formmatting
    blue_fill = op.styles.PatternFill(start_color="00aadd", fill_type='solid')

    # define border style
    thin_border = op.styles.borders.Border(left=op.styles.Side(style='thin'), right=op.styles.Side(style='thin'), top=op.styles.Side(style='thin'), bottom=op.styles.Side(style='thin'))

    # define Text wrap
    text_wrap = op.styles.Alignment(wrap_text=True)

    # Format the header row
    for row in range(1, 2):  # Loop only the 1st row
        for col in range(1, ws.max_column + 1):  # loop through all columns
            ws.cell(row=row, column=col).fill = blue_fill

    # Format all cells
    for row in ws.iter_rows():
        for cell in row:
            # Draw borders
            cell.border = thin_border
            # Wrap all columns
            cell.alignment = text_wrap
    # Save back as same file name
    wb.save(xl) 

def main(url):
    """[summary]

    Args:
        url ([type]): [description]
    """
    global parent_url
    global xl_name
    global domain
    global sheet_name
    parent_url = url
    domain = urlparse(parent_url).netloc
    xl_name = os.path.join(EXCEL_DIR, get_filename(url) + EXCEL_EXTENSION)
    sheet_name = "URLs"
    options = Options()
    options.headless = True
    driver = webdriver.Chrome(executable_path=CHROME_DRIVER, options=options)

    url_list = list()
    crawled_urls = list()

    url_list.append(parent_url)

    # Initiate the crawling by passind the beginning url
    crawled_urls, url_list = crawl_urls(url_list, crawled_urls, driver, parent_url)
    df = pd.DataFrame()
    df['PAGE_TITLE'] = [get_page_title(u) for u in url_list]
    df['URL'] = url_list
    df.to_csv(xl_name.replace(EXCEL_EXTENSION, CSV_EXTENSION), index=False)
    # Finally quit the browser
    driver.quit()

    print ("FULL URLs LIST")
    print (len(set(url_list)))

    print ("CRAWLED URLs LIST")
    print (len(set(crawled_urls)))

    # Load the match list to excel
    load_to_excel(url_list)

    # Format the excel file
    format_excel(xl_name, sheet_name)

if __name__ == "__main__":
    url = "http://localhost:4000"
    main(url)